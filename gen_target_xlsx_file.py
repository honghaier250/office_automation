import pandas as pd
import openpyxl
from openpyxl import load_workbook

def process_midautumn_data_individual():
    try:
        # 1. 从人员名单中提取姓名列表
        source_file = "人员名单.xlsx"
        source_df = pd.read_excel(source_file)
        names_list = source_df['姓名'].tolist()
        print(f"成功从'{source_file}'读取{len(names_list)}个姓名")

        # 2. 将每个姓名依次填写在发放清单预期位置，生成包含姓名、卡号、密码的独立文件
        target_file = "卡号清单.xlsx"

        for i, name in enumerate(names_list):
            print(f"\n正在处理第{i+1}个人员: {name}")

            # 每次处理都需要重新加载工作簿以确保模板状态
            wb = load_workbook(target_file)
            ws = wb.active

            # 将姓名写入目标单元格
            ws['L8'] = name

            # 保存到独立文件
            temp_file = f"xlsx/{name}.xlsx"
            wb.save(temp_file)
            wb.close()

        print(f"\n所有{len(names_list)}个人员处理完成！")

    except FileNotFoundError as e:
        print(f"文件未找到: {e}")
    except KeyError as e:
        print(f"列名可能不存在，请确保Excel中存在'姓名'列: {e}")
    except Exception as e:
        print(f"处理过程中发生错误: {e}")

if __name__ == "__main__":
    process_midautumn_data_individual()
